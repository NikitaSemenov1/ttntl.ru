from openpyxl import load_workbook

wb = load_workbook('../files/tt.xlsx')

sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

date = sheet['A1'].value

group_points = ['A2', 'F2', 'K2', 'P2', 'U2', 'Z2', 'AE2',
          'A13', 'F13', 'K13', 'P13', 'U13', 'Z13']

groups = [str(sheet[point].value) for point in group_points]

def parse_group(point):
    x = sheet[point].row
    y = sheet[point].column
    x += 1
    
    group_tt = []
    col_y = [0, 2, 4]
    for x_d in range(10):
        if x_d == 6:
            continue
        col = []
        for y_d in col_y:
            item = str(sheet.cell(row=x+x_d, column=y+y_d).value)
            if item == 'None':
                item = ""
            if y_d != 4 and sheet.cell(row=x+x_d, column=y+y_d+1).value != "" and sheet.cell(row=x+x_d, column=y+y_d+1).value != None:
                item += " | " + str(sheet.cell(row=x+x_d, column=y+y_d+1).value)
            col.append(item)
        group_tt.append(col)
        
    return group_tt

def parse():
    tt = {}
    tt['date'] = date
    for group_point in group_points:
        tt[str(sheet[group_point].value)] = parse_group(group_point)
    return tt
